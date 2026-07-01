"""MM-TBA -- Multi-Modal Teaching Behavior Analysis.

Target: mean of first 4 GPT-4 rubric scores (continuous).
Features: 13 items from transcript text analysis + metadata.
Groups: None (no grouping metadata).
"""

import numpy as np
import re, io, zipfile, urllib.request
from pathlib import Path

from . import DatasetSource, register_source

FEATURE_NAMES = [
    "char_count", "sentence_count", "avg_sentence_length",
    "question_count", "exclamation_count", "pause_marker_count",
    "digit_count", "math_keyword_count", "gender", "qualification",
    "experience_years", "grade_index", "subject_index",
]
MATH_KEYWORDS = ["函数", "方程", "不等式", "角", "圆", "分数", "概率", "几何", "证明"]
GRADE_MAP = {"p1":1,"p2":2,"p3":3,"p4":4,"p5":5,"p6":6,"j1":7,"j2":8,"j3":9,"s1":10,"s2":11,"s3":12}
SUBJECT_MAP = {"math":1,"mathematics":1,"english":2,"chinese":3,"physics":4,"chemistry":5,"biology":6,"history":7,"geography":8,"politics":9}


@register_source
class MMTBASource(DatasetSource):
    name = "mm_tba"
    display_name = "MM-TBA Teaching Behavior Analysis"
    version = "1.0"
    source_url = "https://github.com/broadsense/MM-TBA"
    license_info = "TBD"
    reference = "Huang et al. (2025)"
    task = "regression"
    n_samples = 186
    n_features = 13
    n_groups = 0
    grouping_description = "None (no grouping metadata)"
    notes = "LLM-scored benchmark. Features extracted from lecture transcripts and GPT-4 rubric reports."

    def download(self):
        import urllib.request, zipfile, io, shutil
        dest_dir = self._ensure_cache_dir()
        meta_path = dest_dir / "metadata.xlsx"
        if meta_path.exists():
            return dest_dir
        # Download repo zip from GitHub
        zip_url = "https://github.com/broadsense/MM-TBA/archive/refs/heads/main.zip"
        resp = urllib.request.urlopen(zip_url, timeout=120)
        with zipfile.ZipFile(io.BytesIO(resp.read())) as z:
            for member in z.namelist():
                target = dest_dir / "/".join(member.split("/")[1:])
                if not target.name:
                    continue
                target.parent.mkdir(parents=True, exist_ok=True)
                with z.open(member) as src, open(target, "wb") as dst:
                    shutil.copyfileobj(src, dst)
        return dest_dir

    def prepare(self):
        import openpyxl

        root = self.download()
        lec_root = root / "Teacher_Lecture_Evaluation"
        meta_path = root / "metadata.xlsx"

        # Load metadata
        metadata_map = {}
        wb = openpyxl.load_workbook(meta_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c or "").strip() for c in rows[0]]
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            rec = dict(zip(header, row))
            sid = str(rec.get("Filename", "")).strip()
            if sid:
                metadata_map[sid] = rec

        # Build samples
        transcript_dir = lec_root / "teacher_lecture_texts"
        X_list, y_list = [], []
        for report_dir in [lec_root / "gpt_report" / "train", lec_root / "gpt_report" / "eval"]:
            if not report_dir.exists():
                continue
            for rp in sorted(report_dir.glob("*.txt")):
                sid = rp.stem
                tp = transcript_dir / f"{sid}.txt"
                if not tp.exists():
                    continue
                rt = rp.read_text(encoding="utf-8", errors="ignore")
                scores = _extract_scores(rt)
                if len(scores) < 4:
                    continue
                tt = tp.read_text(encoding="utf-8", errors="ignore")
                meta = metadata_map.get(sid, {})
                X_list.append(_text_features(tt, meta))
                y_list.append(float(np.mean(scores[:4])))

        X = np.array(X_list, dtype=float)
        y = np.array(y_list, dtype=float)
        card = {"n_samples": len(y), "n_features": X.shape[1],
                "n_groups": 0, "source": "MM-TBA (GitHub)",
                "features": FEATURE_NAMES}
        return X, y, None, card


# ---- inline preprocessing functions ----

def _extract_scores(text):
    tokens = re.findall(r"分数[:：]\s*([0-9]+(?:\.[0-9]+)?(?:\s*[~-]\s*[0-9]+(?:\.[0-9]+)?)?)", text)
    scores = []
    for tok in tokens:
        try:
            scores.append(_parse_score(tok))
        except ValueError:
            continue
    return scores

def _parse_score(tok):
    tok = tok.replace(" ", "")
    if "~" in tok:
        l, r = tok.split("~", 1)
        return (float(l) + float(r)) / 2
    if "-" in tok:
        l, r = tok.split("-", 1)
        return (float(l) + float(r)) / 2
    return float(tok)

def _split_sentences(text):
    return [s for s in re.split(r"[。！？!?]+", text) if s.strip()]

def _parse_experience(value):
    if value is None:
        return 0.0
    text = str(value).strip().lower()
    if not text:
        return 0.0
    m = re.search(r"(\d+(?:\.\d+)?)\s*([a-z]+)", text)
    if not m:
        try:
            return float(text)
        except ValueError:
            return 0.0
    amount = float(m.group(1))
    return amount / 12.0 if m.group(2).startswith("m") else amount

def _text_features(transcript, meta):
    transcript = transcript.strip()
    sentences = _split_sentences(transcript)
    char_count = len(transcript)
    sentence_count = len(sentences)
    avg_sl = char_count / sentence_count if sentence_count else 0.0
    return [
        float(char_count), float(sentence_count), avg_sl,
        float(transcript.count("？") + transcript.count("?")),
        float(transcript.count("！") + transcript.count("!")),
        float(transcript.count("嗯") + transcript.count("啊") + transcript.count("呃")),
        float(sum(ch.isdigit() for ch in transcript)),
        float(sum(transcript.count(kw) for kw in MATH_KEYWORDS)),
        float(meta.get("Gender") or 0),
        float(meta.get("Teacher Qualification Certificate") or 0),
        _parse_experience(meta.get("Teaching Experience")),
        float(GRADE_MAP.get(str(meta.get("Grade") or "").strip().lower(), 0)),
        float(SUBJECT_MAP.get(str(meta.get("Subject") or "").strip().lower(), 0)),
    ]
